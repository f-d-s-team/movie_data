# 导入相关库
import requests
from lxml import etree
from openpyxl import workbook  # 写入Excel表所用
from openpyxl import load_workbook  # 读取Excel表所用
from bs4 import BeautifulSoup as bs
import os
print(os.getcwd())
#os.chdir('C:\Users\19652\Desktop')  # 更改工作目录为桌面
fb=open('豆瓣电影.txt','w',encoding='utf-8')
# 1.将目标网页的内容请求下来
headers = {
    "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/68.0.3440.106 Safari/537.36",
    "Referer": "https://www.douban.com/"
}
douban_url = "https://movie.douban.com/cinema/nowplaying/xian/"
response = requests.get(douban_url, headers=headers)
douban_text = response.text

# 2.将抓取的数据进行处理
html_element = etree.HTML(douban_text)
ul = html_element.xpath('//ul[@class="lists"]')[0]
lis = ul.xpath('./li')
movies = []
titles=[]
scores=[]
stars=[]
durations=[]
regions=[]
directors=[]
actorss=[]
posts=[]
movie1s=[]
for li in lis:
    title = li.xpath('./@data-title')[0]
    score = li.xpath('./@data-score')[0]
    star = li.xpath('./@data-star')[0]
    duration = li.xpath('./@data-duration')[0]
    region = li.xpath('./@data-region')[0]
    director = li.xpath('./@data-director')[0]
    actors = li.xpath('./@data-actors')[0]
    post = li.xpath('.//img/@src')[0]
    movie = {
        "title": title,
        "score": score,
        "star": star,
        "duration": duration,
        "region": region,
        "director": director,
        "actors": actors,
        "post": post
    }
    movie1={
        title,
        score,
        star,
        duration,
        region,
        director,
        actors,
        post
    }
    titles.append(title)
    scores.append(score)
    stars.append(star)
    durations.append(duration)
    regions.append(region )
    directors.append(director)
    actorss.append(actors)
    posts.append(post)
    movies.append(movie)
    movie1s.append(movie1)
for movie in movies:
    print(movie)
print(titles)
print(scores)
print(stars)
print(durations)
print(regions)
print(directors)
print(actorss)
print(posts)
print(len(titles))
#fb=open('豆瓣电影.xlsx','w',encoding='utf-8')
#ws.append(['电影名', '评分', '五角星', '时长','国家/地区','导演', '主演', '海报'])

    #   读取存在的Excel表测试
    #     wb = load_workbook('test.xlsx') #加载存在的Excel表
    #     a_sheet = wb.get_sheet_by_name('Sheet1') #根据表名获取表对象
    #     for row in a_sheet.rows: #遍历输出行数据
    #         for cell in row: #每行的每一个单元格
    #             print cell.value,

    #  创建Excel表并写入数据
import xlwt
book = xlwt.Workbook(encoding='utf-8',style_compression=0)
sheet = book.add_sheet('mysheet',cell_overwrite_ok=True)
movies_info=["电影名", "评分", "五角星","时长","国家/地区","导演", "主演", "海报"]
for j in range(8):
    sheet.write(0,j,movies_info[j])

for i in range(len(titles)):
    sheet.write(i+1,0,titles[i])
    sheet.write(i+1,1,scores[i])
    sheet.write(i+1,2,stars[i])
    sheet.write(i+1,3,durations[i])
    sheet.write(i + 1, 4, regions[i])
    sheet.write(i + 1, 5, directors[i])
    sheet.write(i + 1, 6, actorss[i])
    sheet.write(i + 1, 7, posts[i])
print("结束")
book.save('豆瓣电影.xls')



